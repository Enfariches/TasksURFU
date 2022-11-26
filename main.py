import re
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import BUILTIN_FORMATS
from datetime import datetime
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit

dic_naming = {'name': 'Название',
              'description': 'Описание',
              'key_skills': 'Навыки',
              'experience_id': 'Опыт работы',
              'premium': 'Премиум-вакансия',
              'employer_name': 'Компания',
              'salary_from': 'Нижняя граница вилки оклада',
              'salary_to': 'Верхняя граница вилки оклада',
              'salary_gross': 'Оклад указан до вычета налогов',
              'salary_currency': 'Идентификатор валюты оклада',
              'area_name': 'Название региона',
              'published_at': 'Дата публикации вакансии'}

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}
heads1 = ["Год", "Средняя зарплата", "Средняя зарплата - ", "Количество вакансий", "Количество вакансий - "]
heads2 = ["Город", "Уровень зарплаты",'', "Город", "Доля вакансий"]


class Report:
    def __init__(self):
        self.report = Report.generate_excel(result, options.parameter[1])

    @staticmethod
    def generate_excel(result, vacancy):
        def as_text(val):
            if val is None:
                return ""
            return str(val)

        def cell_parameters(sheet):
            thin = Side(border_style="thin", color="000000")
            for column in sheet.columns:
                length = max(len(as_text(cell.value)) for cell in column)
                sheet.column_dimensions[column[0].column_letter].width = length + 2
                for cell in column:
                    if isinstance(cell.value, float):
                        cell.number_format = BUILTIN_FORMATS[10]
                    if cell.column_letter != "C" and sheet == sheet2:
                        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    if sheet == sheet1:
                        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            return sheet


        salary_by_years, vac_salary_by_years, vacs_by_years, vac_counts_by_years, salary_by_cities, vacs_by_cities = result
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = "Статистика по годам"
        sheet2 = wb.create_sheet("Статистика по городам")
        heads3 = [s.replace('-', f'- {vacancy}') for s in heads1]
        for i, value in enumerate(heads3):
            sheet1.cell(row=1, column=(i+1), value=value).font = Font(bold=True)
        for key, value in salary_by_years.items():
            sheet1.append([key, value, vac_salary_by_years[key], vacs_by_years[key], vac_counts_by_years[key]])
        for i, value in enumerate(heads2):
            sheet2.cell(row=1, column=(i+1), value=value).font = Font(bold=True)
        for (key, value), (k, v) in zip(salary_by_cities.items(), vacs_by_cities.items()):
            sheet2.append([key, value, '', k, v])

        cell_parameters(sheet1)
        cell_parameters(sheet2)

        return wb.save("report.xlsx")
    @staticmethod
    def Graphics(result, vacancy):

        def slash(citites):
            citites = [s.replace('-', '\n').replace(' ', '\n') for s in citites]
            return citites

        def top10(dict):
            first10pairs = {k: dict[k] for k in list(dict)[:11]}
            lastpairs = {k: dict[k] for k in list(dict)[10:]}
            count = 0
            for i in lastpairs.values():
                count += i
            lastpairscount = {"Другие": count}
            first10pairs.update(lastpairscount)
            return first10pairs

        salary_by_years, vac_salary_by_years, vacs_by_years, vac_counts_by_years, salary_by_cities, vacs_by_cities = result
        width = 0.4
        x_nums = np.arange(len(salary_by_years.keys()))
        x_list1 = list(map(lambda x: x - width / 2, x_nums))
        x_list2 = list(map(lambda x: x + width / 2, x_nums))


        fig = plt.figure()
        ax = fig.add_subplot(221)
        ax.set_title("Уровень зарплат по годам")
        ax.bar(x_list1, salary_by_years.values(), width, label='средняя з/п')
        ax.bar(x_list2, vac_salary_by_years.values(), width, label=f'з/п {vacancy}')
        ax.set_xticks(x_nums, salary_by_years.keys(), rotation="vertical")
        ax.legend(loc='upper left', fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis="y")

        ax = fig.add_subplot(222)
        ax.set_title("Количество вакансий по годам")
        ax.bar(x_list1, vacs_by_years.values(), width, label='количество вакансий')
        ax.bar(x_list2, vac_counts_by_years.values(), width, label=f'количество вакансий \n{vacancy}')
        ax.set_xticks(x_nums, vacs_by_years.keys(), rotation="vertical")
        ax.legend(loc='upper left', fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis="y")

        ax = fig.add_subplot(223)
        ax.set_title("Уровень зарплат по городам")
        ax.barh(list(reversed(slash(salary_by_cities.keys()))), list(reversed(list(salary_by_cities.values()))))
        plt.yticks(fontsize=6,linespacing=0.66)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis="x")

        ax = fig.add_subplot(224)
        cities_finaly = top10(vacs_by_cities)
        ax.set_title("Доля вакансий по городам")
        ax.pie(list(cities_finaly.values()), labels=list(cities_finaly.keys()), textprops={'fontsize':6})
        ax.axis("equal")
        plt.tight_layout()

        return plt.savefig('graph.png', dpi=300)

    def generate_pdf(result ,vacancy, heads1, heads2):
        salary_by_years, vac_salary_by_years, vacs_by_years, vac_counts_by_years, salary_by_cities, vacs_by_cities = result
        config = pdfkit.configuration(wkhtmltopdf=r'E:\apps\wkhtmltopdf\bin\wkhtmltopdf.exe')

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        pdf_template = template.render({'vacancy': vacancy, "heads1": heads1, "salary_by_years": salary_by_years,
                                        "vac_salary_by_years": vac_salary_by_years, "vacs_by_years": vacs_by_years,
                                        "vac_counts_by_years": vac_counts_by_years, "heads2": heads2,
                                        "salary_by_cities": salary_by_cities, "vacs_by_cities": vacs_by_cities})

        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


class Interface:
    def __init__(self):
        self.parameter = Interface.get_parameters()

    @staticmethod
    def get_parameters():
<<<<<<< HEAD
        file_name = input("Введите название files: ")
=======
        file_name = input("Введите название чего-то там: ")
>>>>>>> 95b5383eff33668f2a480877a690eeb536039953
        vacancy = input("Введите название профессии: ")
        method = input("Вакансии или Статистика: ")
        return file_name, vacancy, method

    @staticmethod
    def printing_data(dic_vacancies, vac_name, method):
        years = set()
        for vacancy in dic_vacancies:
            years.add(int(datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y')))
        years = sorted(list(years))
        years = list(range(min(years), max(years) + 1))

        salary_by_years = {year: [] for year in years}
        vac_salary_by_years = {year: [] for year in years}

        vacs_by_years = {year: 0 for year in years}
        vac_counts_by_years = {year: 0 for year in years}

        for vacancy in dic_vacancies:
            year = int(datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y'))
            salary_by_years[year].append(vacancy.salary.get_salary_rubles())
            vacs_by_years[year] += 1
            if vac_name in vacancy.name:
                vac_salary_by_years[year].append(vacancy.salary.get_salary_rubles())
                vac_counts_by_years[year] += 1

        salary_by_years = {key: int(sum(value) / len(value)) if len(value) != 0 else 0 for key, value in
                           salary_by_years.items()}
        vac_salary_by_years = {key: int(sum(value) / len(value)) if len(value) != 0 else 0 for key, value in
                               vac_salary_by_years.items()}

        dic_area_name = {}
        for vacancy in dic_vacancies:
            if vacancy.area_name not in dic_area_name:
                dic_area_name[vacancy.area_name] = [vacancy.salary.get_salary_rubles()]
            else:
                dic_area_name[vacancy.area_name].append(vacancy.salary.get_salary_rubles())

        area_name_list = dic_area_name.items()
        area_name_list = [x for x in area_name_list if len(x[1]) / len(dic_vacancies) > 0.01]
        area_name_list = sorted(area_name_list, key=lambda item: sum(item[1]) / len(item[1]), reverse=True)
        salary_by_cities = {item[0]: int(sum(item[1]) / len(item[1])) for item in area_name_list[0: min(len(area_name_list), 10)]}
        vacs_dic = {}

        for vacancy in dic_vacancies:
            if vacancy.area_name in vacs_dic:
                vacs_dic[vacancy.area_name] += 1
            else:
                vacs_dic[vacancy.area_name] = 1

        vacs_counts = {x: round(y / len(dic_vacancies), 4) for x, y in vacs_dic.items()}
        vacs_counts = {k: val for k, val in vacs_counts.items() if val >= 0.01}
        vacs_by_cities = dict(sorted(vacs_counts.items(), key=lambda item: item[1], reverse=True))
        vacs_by_cities = dict(list(vacs_by_cities.items())[:10])
        if method == "Вакансии":
            print("Динамика уровня зарплат по годам:", salary_by_years)
            print("Динамика количества вакансий по годам:", vacs_by_years)
            print("Динамика уровня зарплат по годам для выбранной профессии:", vac_salary_by_years)
            print("Динамика количества вакансий по годам для выбранной профессии:", vac_counts_by_years)
            print("Уровень зарплат по городам (в порядке убывания):", salary_by_cities)
            print("Доля вакансий по городам (в порядке убывания):", vacs_by_cities)
            exit()
        elif method == "Статистика":
            return salary_by_years, vac_salary_by_years, vacs_by_years, vac_counts_by_years, salary_by_cities, vacs_by_cities


class DataSet:
    def __init__(self, file_name):
        self.file_name = file_name
        self.vacancies_objects = DataSet.csv_filter(file_name)

    @staticmethod
    def cleaner_string(text):
        text = re.sub(r"<[^>]+>", "", text)
        text = " ".join(text.split())
        return text

    @staticmethod
    def csv_filter(file_name):
        list_naming, vacancies = DataSet.csv_reader(file_name)
        okay = [x for x in vacancies if len(x) == len(list_naming) and '' not in x]
        people_data = []
        dic_changed_vacancies = {}
        for line in okay:
            for i in range(len(line)):
                dic_changed_vacancies[list_naming[i]] = DataSet.cleaner_string(line[i])
            people_data.append(Vacancy(dic_changed_vacancies['name'],
                                Salary(dic_changed_vacancies['salary_from'], dic_changed_vacancies['salary_to'],
                                dic_changed_vacancies['salary_currency']),
                                dic_changed_vacancies['area_name'], dic_changed_vacancies['published_at']))
        return people_data

    @staticmethod
    def csv_reader(file_name):
       with open(file_name, encoding="utf_8_sig") as file:
            text = csv.reader(file)
            data = [x for x in text]
            if len(data) == 0:
                print("Пустой файл")
                exit()
            list_naming = data[0]
            vacancies = data[1:]
            return list_naming, vacancies

    @staticmethod
    def test_data(arg,method):
        if arg is not None:
            data = DataSet(arg[0])
            return Interface.printing_data(data.vacancies_objects, arg[1], method)


class Salary:
    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.salary_rubles = int((float(self.salary_from) + float(self.salary_to)) / 2) \
                             * currency_to_rub[self.salary_currency]

    def get_salary_rubles(self):
        return self.salary_rubles

class Vacancy:
    def __init__(self, name, salary, area_name, published_at):
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at

options = Interface()
result = DataSet.test_data(options.parameter, options.parameter[2])
Report.Graphics(result, options.parameter[1])
Report.generate_excel(result, options.parameter[1])
Report.generate_pdf(result, options.parameter[1], heads1, heads2)
